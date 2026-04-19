"""Тесты экспорта отчётов в .docx/.xlsx."""

from __future__ import annotations

from email_analyzer.utils.reports import export_emails_to_docx, export_emails_to_xlsx


SAMPLE = [
    {
        "sender": "a@b.com",
        "recipient": "me@me.com",
        "subject": "Test 1",
        "sent_at": "2026-04-19",
        "category": "work",
        "confidence": 0.9,
        "is_spam": False,
        "is_phishing": False,
    },
    {
        "sender": "spam@shop.com",
        "recipient": "me@me.com",
        "subject": "Скидка 70%",
        "sent_at": "2026-04-18",
        "category": "promo",
        "confidence": 0.7,
        "is_spam": True,
        "is_phishing": False,
    },
]


def test_export_docx_creates_file(tmp_path):
    target = tmp_path / "report.docx"
    path = export_emails_to_docx(SAMPLE, target)
    assert path.exists()
    assert path.stat().st_size > 0


def test_export_xlsx_creates_file(tmp_path):
    from openpyxl import load_workbook

    target = tmp_path / "report.xlsx"
    path = export_emails_to_xlsx(SAMPLE, target)
    assert path.exists()

    wb = load_workbook(path)
    ws = wb.active
    assert ws.cell(row=1, column=1).value == "От"
    assert ws.cell(row=2, column=1).value == "a@b.com"
    assert ws.cell(row=3, column=7).value == "да"  # is_spam
